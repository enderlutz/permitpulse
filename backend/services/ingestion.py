"""Parse Houston Public Works weekly Permit eReport xlsx files into Permit rows.

File format (verified 2026-05-21 against Web-eReport-Permits-11-24-2025.xlsx):
    rows 0..2  metadata (title, From, To)
    row 3      blank
    row 4      headers: [Zip Code, Permit Date, Permit Type, Project No, Address, Comments]
    row 5+     data
"""
import re
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


# Use class heuristics from comments
USE_CLASS_PATTERNS = [
    ("warehouse", re.compile(r"\b(warehouse|industrial|distribution)\b", re.I)),
    ("retail", re.compile(r"\b(retail|shopping|store|mall|plaza)\b", re.I)),
    ("restaurant", re.compile(r"\b(restaurant|coffee|cafe|bar|kitchen)\b", re.I)),
    ("office", re.compile(r"\b(office|professional)\b", re.I)),
    ("apartment", re.compile(r"\b(apartment|multifamily|multi[- ]family)\b", re.I)),
    ("residential", re.compile(r"\b(single[- ]family|sfr|residence|dwelling|new home)\b", re.I)),
]

SQ_FT_RE = re.compile(r"([\d,]+)\s*SF\b", re.I)


def classify(comments: str | None) -> str | None:
    if not comments:
        return None
    for label, pat in USE_CLASS_PATTERNS:
        if pat.search(comments):
            return label
    return None


def extract_sq_ft(comments: str | None) -> int | None:
    if not comments:
        return None
    m = SQ_FT_RE.search(comments)
    if not m:
        return None
    try:
        return int(m.group(1).replace(",", ""))
    except ValueError:
        return None


def parse_xlsx(path: str | Path) -> Iterable[dict]:
    """Yield permit dicts from a single weekly eReport xlsx."""
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 5:
            continue
        if not row or row[0] is None:
            continue
        zip_code, permit_date, permit_type, project_no, address, comments = (
            row[0],
            row[1],
            row[2],
            row[3],
            row[4],
            row[5] if len(row) > 5 else None,
        )
        if isinstance(permit_date, datetime):
            permit_date = permit_date.date()
        elif isinstance(permit_date, str):
            try:
                permit_date = datetime.fromisoformat(permit_date).date()
            except Exception:
                permit_date = None

        yield {
            "project_no": str(project_no) if project_no is not None else None,
            "permit_date": permit_date,
            "permit_type": str(permit_type).strip() if permit_type else None,
            "address": str(address).strip() if address else None,
            "zip_code": str(zip_code).strip() if zip_code else None,
            "comments": str(comments).strip() if comments else None,
            "use_class": classify(str(comments) if comments else None),
            "square_feet": extract_sq_ft(str(comments) if comments else None),
        }
